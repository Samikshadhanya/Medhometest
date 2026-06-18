import asyncio
import time
import json
import sys
from pathlib import Path
from datetime import datetime
import aiohttp
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Force UTF-8 output on Windows to handle special chars
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE_DIR = Path(__file__).parent
ROOT_DIR = BASE_DIR.parent.parent
INPUT_FILE = ROOT_DIR / "input.json"

# Load config
if INPUT_FILE.exists():
    with open(INPUT_FILE, "r") as f:
        CFG = json.load(f)
    BASE_URL = CFG["baseUrl"].rstrip("/")
else:
    BASE_URL = "https://medhome-81dq.vercel.app"

ENDPOINTS = [
    ("/", "Home Page"),
    ("/signin", "Sign In Page"),
]

VIRTUAL_USERS = 100
DURATION_SECONDS = 60

# Statistics lists
results = []  # will contain dicts: {"endpoint": str, "status": int, "latency_ms": float, "success": bool}

async def worker(session, worker_id, stop_event):
    """Simulates a single virtual user making continuous requests."""
    import random
    while not stop_event.is_set():
        path, name = random.choice(ENDPOINTS)
        url = BASE_URL + path
        t0 = time.time()
        try:
            async with session.get(url, timeout=10, allow_redirects=True) as resp:
                status = resp.status
                latency = (time.time() - t0) * 1000
                success = status in [200, 301, 302, 307, 308]
                results.append({
                    "endpoint": path,
                    "name": name,
                    "status": status,
                    "latency_ms": latency,
                    "success": success
                })
        except Exception as e:
            latency = (time.time() - t0) * 1000
            results.append({
                "endpoint": path,
                "name": name,
                "status": 0,
                "latency_ms": latency,
                "success": False,
                "error": str(e)
            })
        
        # Pacing: small sleep to mimic natural user wait, but keeping it small to generate high load
        await asyncio.sleep(random.uniform(0.05, 0.15))

async def run_load_test():
    print(f"======================================================")
    print(f"Starting MedHome Load Testing Suite...")
    print(f"Target URL     : {BASE_URL}")
    print(f"Virtual Users  : {VIRTUAL_USERS}")
    print(f"Duration       : {DURATION_SECONDS} seconds")
    print(f"======================================================")

    stop_event = asyncio.Event()
    
    # Configure connection pool for high concurrency
    connector = aiohttp.TCPConnector(limit=VIRTUAL_USERS, ssl=False)
    
    async with aiohttp.ClientSession(connector=connector) as session:
        # Spawn virtual users
        tasks = [asyncio.create_task(worker(session, i, stop_event)) for i in range(VIRTUAL_USERS)]
        
        # Run for the specified duration
        start_time = time.time()
        print("Warmup complete. Load test is running...")
        
        # Simple countdown progress updates
        for elapsed in range(1, DURATION_SECONDS + 1):
            await asyncio.sleep(1)
            if elapsed % 10 == 0:
                print(f"  Progress: {elapsed}/{DURATION_SECONDS}s elapsed | Requests sent so far: {len(results)}")
        
        # Signal workers to stop
        stop_event.set()
        
        # Wait for all workers to finish current request
        await asyncio.gather(*tasks, return_exceptions=True)
        
    actual_duration = time.time() - start_time
    generate_reports(actual_duration)

def generate_reports(duration):
    total_requests = len(results)
    if total_requests == 0:
        print("Error: No requests were sent during the load test.")
        return

    successes = [r for r in results if r["success"]]
    failures = [r for r in results if not r["success"]]
    latencies = [r["latency_ms"] for r in results]

    avg_latency = sum(latencies) / total_requests
    min_latency = min(latencies)
    max_latency = max(latencies)
    rps = total_requests / duration

    # Percentiles
    sorted_latencies = sorted(latencies)
    p95 = sorted_latencies[int(total_requests * 0.95)] if total_requests > 0 else 0
    p99 = sorted_latencies[int(total_requests * 0.99)] if total_requests > 0 else 0

    print(f"\n======================================================")
    print(f"Load Test Execution Complete")
    print(f"======================================================")
    print(f"Actual Duration     : {duration:.2f} s")
    print(f"Total Requests      : {total_requests}")
    print(f"Successful Requests : {len(successes)} ({len(successes)/total_requests*100:.1f}%)")
    print(f"Failed Requests     : {len(failures)} ({len(failures)/total_requests*100:.1f}%)")
    print(f"Requests Per Second : {rps:.2f} req/sec")
    print(f"Latency (Response Time):")
    print(f"  Min               : {min_latency:.1f} ms")
    print(f"  Avg               : {avg_latency:.1f} ms")
    print(f"  Max               : {max_latency:.1f} ms")
    print(f"  95th Percentile   : {p95:.1f} ms")
    print(f"  99th Percentile   : {p99:.1f} ms")
    print(f"======================================================")

    # Save to JSON
    report_json = BASE_DIR / "load_report.json"
    with open(report_json, "w") as f:
        json.dump({
            "target_url": BASE_URL,
            "virtual_users": VIRTUAL_USERS,
            "duration_seconds": duration,
            "total_requests": total_requests,
            "successful_requests": len(successes),
            "failed_requests": len(failures),
            "requests_per_second": rps,
            "min_latency_ms": min_latency,
            "avg_latency_ms": avg_latency,
            "max_latency_ms": max_latency,
            "p95_latency_ms": p95,
            "p99_latency_ms": p99,
        }, f, indent=2)

    # Generate Excel Report
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Load Test Report"

    # Formatting styles
    bold = Font(bold=True)

    # Summary section (Rows 1-9)
    ws.cell(row=1, column=1, value="Load Test Staging Summary").font = Font(bold=True, size=12)
    
    summary_data = [
        ("Target URL", BASE_URL),
        ("Virtual Users", VIRTUAL_USERS),
        ("Test Duration (s)", round(duration, 2)),
        ("Total Requests Sent", total_requests),
        ("Successful Requests", len(successes)),
        ("Failed Requests", len(failures)),
        ("Average RPS (req/sec)", round(rps, 2)),
        ("Average Response Time (ms)", round(avg_latency, 1)),
        ("Min Response Time (ms)", round(min_latency, 1)),
        ("Max Response Time (ms)", round(max_latency, 1)),
        ("95th Percentile (ms)", round(p95, 1)),
        ("99th Percentile (ms)", round(p99, 1)),
    ]

    for idx, (label, val) in enumerate(summary_data, 3):
        ws.cell(row=idx, column=1, value=label).font = bold
        ws.cell(row=idx, column=2, value=val)

    # Page wise performance table
    ws.cell(row=17, column=1, value="Endpoint Performance Breakdown").font = Font(bold=True, size=12)
    
    headers = [
        "Endpoint", "Page Name", "Total Requests", "Successes", "Failures", 
        "Avg RPS", "Avg Latency (ms)", "Min (ms)", "Max (ms)"
    ]
    col_widths = [18, 18, 16, 14, 14, 14, 18, 12, 12]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=19, column=col_idx, value=header)
        cell.font = bold
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Calculate metrics per endpoint
    endpoints_stats = {}
    for ep, name in ENDPOINTS:
        ep_reqs = [r for r in results if r["endpoint"] == ep]
        if ep_reqs:
            ep_succ = [r for r in ep_reqs if r["success"]]
            ep_fail = [r for r in ep_reqs if not r["success"]]
            ep_lats = [r["latency_ms"] for r in ep_reqs]
            
            endpoints_stats[ep] = {
                "name": name,
                "total": len(ep_reqs),
                "success": len(ep_succ),
                "failure": len(ep_fail),
                "rps": len(ep_reqs) / duration,
                "avg": sum(ep_lats) / len(ep_reqs),
                "min": min(ep_lats),
                "max": max(ep_lats)
            }

    row_num = 20
    for ep, stats in endpoints_stats.items():
        row_values = [
            ep,
            stats["name"],
            stats["total"],
            stats["success"],
            stats["failure"],
            round(stats["rps"], 2),
            round(stats["avg"], 1),
            round(stats["min"], 1),
            round(stats["max"], 1)
        ]
        for col_idx, val in enumerate(row_values, 1):
            ws.cell(row=row_num, column=col_idx, value=val)
        row_num += 1

    timestamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    output_filename = BASE_DIR / f"Load_Test_Report_{timestamp}.xlsx"
    wb.save(output_filename)
    print(f"Excel report saved: {output_filename}")

if __name__ == '__main__':
    asyncio.run(run_load_test())
